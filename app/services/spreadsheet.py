from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import customtkinter as ctk

def salvar_dados_excel(data, os_num, servico, valor):
    caminho_arquivo = Path(__file__).parents[2] / "MIMO.xlsx"

    print(f"Usando caminho: {caminho_arquivo}")  # Diagnóstico opcional

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        celula_os = row[0]  # A única célula da coluna C nesta linha
        if celula_os.value and str(celula_os.value).strip() == str(os_num).strip():
            return False

    linha = 2
    while ws[f"A{linha}"].value:
        linha += 1

    ws[f"A{linha}"] = data
    ws[f"C{linha}"] = os_num
    ws[f"D{linha}"] = servico
    ws[f"E{linha}"] = valor

    wb.save(caminho_arquivo)
    return True

class servicos(ctk.CTkScrollableFrame):
    def __init__(self, master):
        super().__init__(master)

        self.label_tabela = ctk.CTkLabel(self, text="", font=("Courier New", 12), anchor="w", justify="left")
        self.label_tabela.grid(row=0, column=0, sticky="w", padx=10, pady=10)

        self.atualizar_dados()  # Exibe os dados ao abrir o app

    def atualizar_dados(self):
        try:
            df = pd.read_excel("MIMO.xlsx", usecols=["DATA", "O.S.", "TIPO MONTAGEM", "VALOR"])
            df.dropna(how='all', inplace=True)
            df.columns = ["Data", "O.S.", "Serviço", "Valor"]

            # Define a largura de cada coluna
            col_widths = {
                "Data": 12,
                "O.S.": 10,
                "Serviço": 20,
                "Valor": 10
            }

            # Formata o cabeçalho
            header = "| " + " | ".join(f"{col:<{col_widths[col]}}" for col in df.columns) + " |"

            # Formata cada linha
            linhas = []
            for _, row in df.iterrows():
                linha = "| " + " | ".join(f"{str(row[col]):<{col_widths[col]}}" for col in df.columns) + " |"
                linhas.append(linha)

            # Junta tudo
            tabela_texto = "\n".join([header] + linhas)

            self.label_tabela.configure(text=tabela_texto)

        except Exception as e:
            self.label_tabela.configure(text=f"Erro ao carregar dados: {e}", text_color="red")
